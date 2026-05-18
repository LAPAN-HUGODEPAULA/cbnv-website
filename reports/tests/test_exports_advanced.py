import pytest
import json
from io import BytesIO

from django.urls import reverse
from accounts.tests.factories import create_user_with_profile
from proceedings.models import FinalMaterial
from submissions.models import Submission, ThematicAxis, SubmissionAuthor
from videos.models import PUBLIC, VideoResource
from openpyxl import load_workbook

@pytest.fixture
def chair_user():
    return create_user_with_profile(
        username="chair_adv", email="chair_adv@test.com", password="pw",
        is_chair=True,
    )

@pytest.mark.django_db
class TestAdvancedExports:
    @pytest.fixture(autouse=True)
    def setup(self, client, chair_user):
        self.client = client
        self.client.login(username="chair_adv", password="pw")
        self.axis = ThematicAxis.objects.create(name="Neuro", order=1)

    def test_xlsx_export_indicators(self):
        resp = self.client.get(reverse("reports:export_indicators"), {"format": "xlsx"})
        assert resp.status_code == 200
        assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        workbook = load_workbook(BytesIO(resp.content))
        assert workbook.sheetnames == ["Indicators"]
        assert [cell.value for cell in workbook.active[1]] == [
            "tipo",
            "categoria",
            "valor",
            "contagem",
        ]

    def test_institutions_export(self):
        # Create some data
        user = create_user_with_profile(username="aut", is_author=True)
        s = Submission.objects.create(
            title="S1", thematic_axis=self.axis, submitter=user
        )
        SubmissionAuthor.objects.create(
            submission=s, first_name="A", last_name="B", email="a@b.com", institution="UFMG"
        )
        
        resp = self.client.get(reverse("reports:export_institutions"), {"format": "csv"})
        assert resp.status_code == 200
        content = resp.content.decode("utf-8-sig")
        assert "UFMG" in content
        assert "1" in content # total submissions or authors

    def test_submission_filters(self):
        user = create_user_with_profile(username="aut2", is_author=True)
        axis2 = ThematicAxis.objects.create(name="Vision", order=2)
        Submission.objects.create(
            title="S_Neuro",
            abstract="Resumo neuro.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
            status="draft",
        )
        Submission.objects.create(
            title="S_Vision",
            abstract="Resumo vision.",
            keywords=["a", "b", "c"],
            thematic_axis=axis2,
            submitter=user,
            status="submitted",
        )
        
        # Test status filter
        resp = self.client.get(reverse("reports:export_submissions"), {"format": "json", "status": "submitted"})
        data = json.loads(resp.content)
        assert len(data) == 1
        assert data[0]["titulo"] == "S_Vision"
        
        # Test thematic axis filter
        resp = self.client.get(reverse("reports:export_submissions"), {"format": "json", "thematic_axis": self.axis.id})
        data = json.loads(resp.content)
        assert len(data) == 1
        assert data[0]["titulo"] == "S_Neuro"

    def test_dashboard_and_indicator_export_apply_modality_filter(self):
        user = create_user_with_profile(username="aut3", is_author=True)
        Submission.objects.create(
            title="Oral",
            abstract="Resumo oral.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
            status="accepted_oral",
            final_modality="oral",
        )
        Submission.objects.create(
            title="Poster",
            abstract="Resumo poster.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
            status="accepted_poster",
            final_modality="poster",
        )

        response = self.client.get(reverse("reports:dashboard"), {"final_modality": "oral"})
        assert response.status_code == 200
        assert response.context["total"] == 1
        assert "final_modality=oral" in response.content.decode()

        response = self.client.get(
            reverse("reports:export_indicators"),
            {"format": "json", "final_modality": "oral"},
        )
        data = json.loads(response.content)
        assert data["resumo_geral"]["total_submissoes"] == 1
        assert data["resumo_geral"]["por_modalidade"] == [
            {"modalidade": "oral", "count": 1}
        ]

    def test_author_and_institution_exports_apply_institution_filter(self):
        user = create_user_with_profile(username="aut4", is_author=True)
        first = Submission.objects.create(
            title="A",
            abstract="Resumo A.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
        )
        second = Submission.objects.create(
            title="B",
            abstract="Resumo B.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
        )
        SubmissionAuthor.objects.create(
            submission=first,
            first_name="Ana",
            last_name="A",
            email="ana@test.com",
            institution="UFMG",
            is_corresponding=True,
        )
        SubmissionAuthor.objects.create(
            submission=second,
            first_name="Bruno",
            last_name="B",
            email="bruno@test.com",
            institution="USP",
        )

        response = self.client.get(
            reverse("reports:export_authors"),
            {"format": "csv", "institution": "UFMG"},
        )
        content = response.content.decode("utf-8-sig")
        assert "Apresentador" in content
        assert "Ana A" in content
        assert "Bruno B" not in content

        response = self.client.get(
            reverse("reports:export_institutions"),
            {"format": "csv", "institution": "UFMG"},
        )
        content = response.content.decode("utf-8-sig")
        assert "UFMG" in content
        assert "USP" not in content

    def test_proceedings_export_fields_and_protected_files_are_safe(self):
        user = create_user_with_profile(username="aut5", is_author=True)
        submission = Submission.objects.create(
            title="Proceedings",
            abstract="Resumo publicado.",
            keywords=["neuro", "visao", "teste"],
            thematic_axis=self.axis,
            submitter=user,
            status="ready_for_proceedings",
            final_modality="oral",
        )
        SubmissionAuthor.objects.create(
            submission=submission,
            first_name="Ana",
            last_name="Publica",
            email="ana.publica@test.com",
            institution="UFMG",
        )
        FinalMaterial.objects.create(
            submission=submission,
            final_pdf="protected/final.pdf",
            publication_authorized=True,
        )

        response = self.client.get(reverse("reports:export_proceedings"), {"format": "csv"})
        content = response.content.decode("utf-8-sig")
        assert "Resumo" in content
        assert "Palavras-chave" in content
        assert "Autorizado" in content
        assert "Resumo publicado." in content
        assert "protected/final.pdf" not in content

    def test_promoted_video_indicator(self):
        user = create_user_with_profile(username="aut6", is_author=True)
        submission = Submission.objects.create(
            title="Video",
            abstract="Resumo video.",
            keywords=["a", "b", "c"],
            thematic_axis=self.axis,
            submitter=user,
            status="accepted_video",
            final_modality="video",
        )
        video = VideoResource.objects.create(
            title="Vídeo público",
            youtube_url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
            status=PUBLIC,
        )
        FinalMaterial.objects.create(
            submission=submission,
            video_url=video.youtube_url,
            video_resource=video,
            publication_authorized=True,
        )

        response = self.client.get(reverse("reports:export_indicators"), {"format": "json"})
        data = json.loads(response.content)
        assert data["materiais"]["with_video"] == 1
        assert data["materiais"]["promoted_videos"] == 1
